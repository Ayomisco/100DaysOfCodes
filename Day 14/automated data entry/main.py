# Importing modules
from fileinput import close
from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

# WINDOWS STARTS HERE
root = Tk()

# Styles
fonts = ("Rockwell", 17)
fg_c_o = "#DE781F"
bg_c = "#655aa2"
fg_c_w = "#ffffff"


root.title("Optimum Data Entry")
root.geometry("720x410+300+200")
root.resizable(0,0)
root.configure(bg=bg_c)
# Icon
icon_file = PhotoImage(file="logo.png")
root.iconphoto(True, icon_file)
# root.overrideredirect(True)

# Using the Pathlib to get the path to save the excel file
file = pathlib.Path('contact_info.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "PhoneNumber"
    sheet["C1"] = "Age"
    sheet["D1"] = "Gender"
    sheet["E1"] = "Address"


    file.save("contact_info.xlsx")

#Defining Functions

def submit():
    # Get the data text inputs
    name = nameValue.get()
    contact = contactValue.get()
    age = ageValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Testing in form of printing
    # print(name)
    # print(contact)
    # print(age)
    # print(gender)
    # print(address)

    #Upddating the excel file
    file = openpyxl.load_workbook("contact_info.xlsx")
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    try:
        file.save("contact_info.xlsx")
        messagebox.showinfo("info", "Detail Add Successfully!")
    except PermissionError:
        messagebox.showwarning("warning", "Please close your workbook application if open!")


    nameValue.set("")
    ageValue.set("")
    contactValue.set("")
    addressEntry.delete(1.0, END)


#Close
def close():
    close_answer = messagebox.askyesno("Application", "Are you sure you wants to close?")
    if close_answer == 1:
        root.destroy()


def clear():
    nameValue.set("")
    ageValue.set("")
    contactValue.set("")
    addressEntry.delete(1.0, END)





#Heading
heading = Label(root, text="Please Fill out this Entry Form: ", font=("Rockwell", 17, "bold"), bg=bg_c, fg=fg_c_w)
heading.place(x=20, y=20) 

#labels
Label(root, text="Name:", font=fonts,bg=bg_c, fg=fg_c_o).place(x=50, y=100)
Label(root, text="Contact No.:", font=fonts,bg=bg_c, fg=fg_c_o).place(x=50, y=150)
Label(root, text="Age:", font=fonts,bg=bg_c, fg=fg_c_o).place(x=50, y=200)
Label(root, text="Gender:", font=fonts,bg=bg_c, fg=fg_c_o).place(x=390, y=200)
Label(root, text="Address:", font=fonts,bg=bg_c, fg=fg_c_o).place(x=50, y=250)


#Entry
nameValue = StringVar()
contactValue = StringVar()
ageValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=35, bd=2, font=fonts)
contactEntry = Entry(root, textvariable=contactValue, width=35, bd=2, font=fonts)
ageEntry = Entry(root, textvariable=ageValue, width=12, bd=2, font=fonts)

# Gender
gender_combobox = Combobox(root, values=['Male', 'Female'], font='Rockwell 14', state='r', width=14, height=47)
gender_combobox.place(x=490, y=200)
gender_combobox.set('Male')

#Address Entry
addressEntry = Text(root, width=60, height=4, bd=4)

nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)
addressEntry.place(x=200, y=250)



# Button
Button(root, text="Submit", bg=bg_c, fg=fg_c_w, width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg=bg_c, fg=fg_c_w, width=15, height=2, command=clear).place(x=360, y=350)
Button(root, text="Exit", bg=bg_c, fg=fg_c_w, width=15, height=2, command=close).place(x=520, y=350)




root.mainloop()